"""Concrete :class:`~core.interfaces.IPDFGenerator` that writes an Excel report.

Produces a ``.xlsx`` file using *openpyxl* with:

* Right-to-left sheet orientation for Hebrew content.
* Styled column-header row (white text on blue background).
* Auto-sized columns based on cell content.
* A metadata block at the top of the sheet (employee, period, totals).

The application layer depends only on :class:`~core.interfaces.IPDFGenerator`;
this module must **never** be imported by the application or domain layers.
Only :class:`~container.Container` wires it in.
"""
from __future__ import annotations

import os

from core.exceptions import GenerationError
from core.interfaces import IPDFGenerator, ILogger
from domain.dtos.report_dtos import ProcessedReportDTO

# Column order for the output sheet (matches the source PDF's logical order)
_OUTPUT_COLUMNS = [
    "תאריך",
    "יום בשבוע",
    "שעות כניסה",
    "שעת יציאה",
    'סה"כ שעות',
    "הערות",
]

_HEADER_BG_COLOR = "4472C4"   # Blue
_HEADER_FONT_COLOR = "FFFFFF"  # White


class ExcelReportGenerator(IPDFGenerator):
    """Serialises a :class:`~domain.dtos.report_dtos.ProcessedReportDTO` to ``.xlsx``.

    Args:
        logger: An :class:`~core.interfaces.ILogger` instance injected by the
            :class:`~container.Container`.

    Example:
        >>> gen = ExcelReportGenerator(logger=my_logger)
        >>> gen.generate(processed_dto, "output/april_2025.xlsx")
    """

    def __init__(self, logger: ILogger) -> None:
        self._logger = logger

    # ------------------------------------------------------------------
    # IPDFGenerator implementation
    # ------------------------------------------------------------------

    def generate(self, data: ProcessedReportDTO, output_path: str) -> None:
        """Write *data* to an Excel workbook at *output_path*.

        Creates any missing parent directories before writing.

        Args:
            data: Fully processed report data to serialise.
            output_path: Destination file path (must end in ``.xlsx``).

        Raises:
            GenerationError: If *openpyxl* is not installed or if the file
                cannot be written (e.g. permission error, disk full).
        """
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError as exc:
            msg = (
                "openpyxl is required for Excel generation. "
                "Install it with: pip install openpyxl"
            )
            self._logger.error(msg)
            raise GenerationError(msg) from exc

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Attendance Report"
            ws.sheet_view.rightToLeft = True  # RTL layout for Hebrew

            bold = Font(bold=True)

            # ----------------------------------------------------------
            # Metadata block
            # ----------------------------------------------------------
            metadata_rows = [
                ("עובד:", data.employee_name or "לא ידוע"),
                ("תקופה:", data.period or "לא ידוע"),
                ("ימי עבודה:", data.working_days),
                ("סה\"כ שעות:", f"{data.total_hours_sum:.2f}"),
            ]
            for label, value in metadata_rows:
                ws.append([label, value])
                ws.cell(row=ws.max_row, column=1).font = bold

            ws.append([])  # blank separator row

            # ----------------------------------------------------------
            # Column header row
            # ----------------------------------------------------------
            ws.append(_OUTPUT_COLUMNS)
            header_row_num = ws.max_row
            header_fill = PatternFill("solid", fgColor=_HEADER_BG_COLOR)
            header_font = Font(bold=True, color=_HEADER_FONT_COLOR)
            center = Alignment(horizontal="center", vertical="center")

            for col_idx in range(1, len(_OUTPUT_COLUMNS) + 1):
                cell = ws.cell(row=header_row_num, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center

            # ----------------------------------------------------------
            # Data rows
            # ----------------------------------------------------------
            for record in data.records:
                ws.append(
                    [
                        record.date,
                        record.day_of_week,
                        record.entry_time,
                        record.exit_time,
                        record.total_hours,
                        record.notes,
                    ]
                )

            # ----------------------------------------------------------
            # Auto-size columns (approximate — openpyxl has no native API)
            # ----------------------------------------------------------
            for col in ws.columns:
                max_length = max(
                    (len(str(cell.value)) for cell in col if cell.value is not None),
                    default=8,
                )
                ws.column_dimensions[col[0].column_letter].width = max(12, max_length + 3)

            # ----------------------------------------------------------
            # Persist
            # ----------------------------------------------------------
            parent_dir = os.path.dirname(os.path.abspath(output_path))
            os.makedirs(parent_dir, exist_ok=True)
            wb.save(output_path)

        except GenerationError:
            raise
        except Exception as exc:
            msg = f"Failed to write Excel report to '{output_path}': {exc}"
            self._logger.error(msg, exc_info=True)
            raise GenerationError(msg) from exc

        self._logger.info(
            f"ExcelReportGenerator: report saved successfully -> '{output_path}'"
        )
